"""
Parser for extracting data from CTCAC Excel applications
"""

import re
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook

from data_models import ApplicationData, ConstructionInterestFees, PermanentFinancing


class ExcelParser:
    """Parser for extracting data from CTCAC Excel applications"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.sources_uses_sheet = None
        self.application_sheet = None

    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.wb = load_workbook(self.file_path, data_only=True)
            # Find Sources & Uses Budget sheet
            for sheet_name in self.wb.sheetnames:
                if "sources" in sheet_name.lower() and "uses" in sheet_name.lower():
                    self.sources_uses_sheet = self.wb[sheet_name]
                elif (
                    sheet_name.lower() == "application"
                    and self.application_sheet is None
                ):
                    self.application_sheet = self.wb[sheet_name]

            if not self.sources_uses_sheet:
                raise ValueError("Could not find 'Sources & Uses Budget' sheet")
        except Exception as e:
            raise ValueError(f"Error loading workbook: {str(e)}")

    def find_cell_by_text(
        self, sheet, search_text: str, exact_match: bool = False
    ) -> Optional[Tuple[int, int]]:
        """Find cell containing search text, returns (row, col) tuple"""
        search_text_lower = search_text.lower()
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value).lower()
                    if exact_match:
                        if cell_text == search_text_lower:
                            return (cell.row, cell.column)
                    else:
                        if search_text_lower in cell_text:
                            return (cell.row, cell.column)
        return None

    def extract_number_from_cell(self, sheet, row: int, col: int) -> float:
        """Extract numeric value from cell, handling formulas and text"""
        cell = sheet.cell(row, col)
        if cell.value is None:
            return 0.0

        # If it's already a number
        if isinstance(cell.value, (int, float)):
            return float(cell.value)

        # If it's a string, try to extract number
        if isinstance(cell.value, str):
            # Remove currency symbols and commas
            cleaned = re.sub(r"[^\d.-]", "", cell.value)
            try:
                return float(cleaned)
            except ValueError:
                return 0.0

        return 0.0

    def find_value_near_label(
        self, sheet, label_text: str, search_range: int = 50
    ) -> Optional[float]:
        """Find a label and extract the value from nearby cells"""
        label_pos = self.find_cell_by_text(sheet, label_text)
        if not label_pos:
            return None

        row, col = label_pos
        # Priority: Check column B first (most common for totals), then C, R, S
        priority_cols = [2, 3, 18, 19]  # B, C, R, S
        for check_col in priority_cols:
            value = self.extract_number_from_cell(sheet, row, check_col)
            if value > 0:
                return value

        # Fallback: Check cells to the right
        for offset in range(1, 10):
            value = self.extract_number_from_cell(sheet, row, col + offset)
            if value > 0:
                return value
        return None

    def extract_construction_interest_fees(self) -> ConstructionInterestFees:
        """Extract Construction Interest & Fees section - fixed order"""
        data = ConstructionInterestFees()
        sheet = self.sources_uses_sheet

        # Find the section header
        section_pos = self.find_cell_by_text(sheet, "CONSTRUCTION INTEREST & FEES")
        if not section_pos:
            return data

        start_row = section_pos[0]
        end_row = start_row + 20  # Look within 20 rows of section start

        # Fixed order: extract line items one by one in order
        # Order: Construction Loan Interest, Origination Fee, Credit Enhancement,
        # Bond Premium, Cost of Issuance, Title & Recording, Taxes, Insurance, Other, Total
        current_row = start_row + 1  # Start after header

        # 1. Construction Loan Interest
        if current_row <= end_row:
            data.construction_loan_interest = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 2. Origination Fee
        if current_row <= end_row:
            data.origination_fee = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 3. Credit Enhancement/Application Fee
        if current_row <= end_row:
            data.credit_enhancement_fee = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 4. Bond Premium
        if current_row <= end_row:
            data.bond_premium = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 5. Cost of Issuance
        if current_row <= end_row:
            data.cost_of_issuance = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 6. Title & Recording
        if current_row <= end_row:
            data.title_recording = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 7. Taxes
        if current_row <= end_row:
            data.taxes = self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            current_row += 1

        # 8. Insurance
        if current_row <= end_row:
            data.insurance = self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            current_row += 1

        # 9. Other (between Insurance and Total)
        if current_row <= end_row:
            full_cell_text = str(sheet.cell(current_row, 1).value or "").strip()
            data.other_description = full_cell_text
            data.other_amount = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 10. Total Construction Interest & Fees
        if current_row <= end_row:
            data.total = self.extract_number_from_cell(sheet, current_row, 2) or 0.0

        return data

    def extract_permanent_financing(self) -> PermanentFinancing:
        """Extract Permanent Financing section - fixed order"""
        data = PermanentFinancing()
        sheet = self.sources_uses_sheet

        # Find the section header
        section_pos = self.find_cell_by_text(sheet, "PERMANENT FINANCING")
        if not section_pos:
            return data

        start_row = section_pos[0]
        end_row = start_row + 15  # Look within 15 rows of section start

        # Fixed order: extract line items one by one in order
        # Order: Loan Origination Fee, Credit Enhancement, Title & Recording,
        # Taxes, Insurance, Other, Total
        current_row = start_row + 1  # Start after header

        # 1. Loan Origination Fee
        if current_row <= end_row:
            data.loan_origination_fee = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 2. Credit Enhancement/Application Fee
        if current_row <= end_row:
            data.credit_enhancement_fee = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 3. Title & Recording
        if current_row <= end_row:
            data.title_recording = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 4. Taxes
        if current_row <= end_row:
            data.taxes = self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            current_row += 1

        # 5. Insurance
        if current_row <= end_row:
            data.insurance = self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            current_row += 1

        # 6. Other (between Insurance and Total)
        if current_row <= end_row:
            full_cell_text = str(sheet.cell(current_row, 1).value or "").strip()
            data.other_description = full_cell_text
            data.other_amount = (
                self.extract_number_from_cell(sheet, current_row, 2) or 0.0
            )
            current_row += 1

        # 7. Total Permanent Financing Costs
        if current_row <= end_row:
            data.total = self.extract_number_from_cell(sheet, current_row, 2) or 0.0

        return data

    def extract_new_construction_total(self) -> Optional[float]:
        """Extract New Construction total from Sources & Uses Budget"""
        sheet = self.sources_uses_sheet

        # Look for "Total New Construction Costs" - search row by row
        for row in range(1, min(sheet.max_row + 1, 100)):
            cell_value = str(sheet.cell(row, 1).value or "").lower()
            if "total new construction costs" in cell_value:
                value = self.extract_number_from_cell(sheet, row, 2)
                if value and value > 0:
                    return value

        # Fallback: look for "NEW CONSTRUCTION" section and find total
        new_const_pos = self.find_cell_by_text(sheet, "NEW CONSTRUCTION")
        if new_const_pos:
            # Look for a total row below this
            for row_offset in range(1, 20):
                row = new_const_pos[0] + row_offset
                cell_value = str(sheet.cell(row, 1).value or "").lower()
                if "total" in cell_value and "construction" in cell_value:
                    value = self.extract_number_from_cell(sheet, row, 2)
                    if value and value > 0:
                        return value

        return None

    def extract_units_and_sf(self) -> Tuple[Optional[int], Optional[float]]:
        """
        Extract total units and square footage from Application tab.

        Dynamically finds labels (tries columns D and G), then extracts values from Column AG (33):
        - "Total number of units"
        - "Total square footage of all project structures"
        """
        if not self.application_sheet:
            return None, None

        sheet = self.application_sheet

        # Try multiple columns for labels (D=4, G=7)
        label_columns = [4, 7]  # Columns D and G
        # Value columns to try: AG (33) first, then L (12) as fallback
        value_columns = [33, 12]  # Columns AG and L

        units: Optional[int] = None
        sf: Optional[float] = None

        # 1. Dynamic search for Total number of units
        # Search for label in multiple columns, extract value from AG or L
        for row in range(1, sheet.max_row + 1):
            for col in label_columns:
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and "total number of units" in str(cell_value).lower():
                    # Skip if it's the "excluding managers" version
                    cell_lower = str(cell_value).lower()
                    if "excluding managers" not in cell_lower:
                        # Try AG first, then L
                        for val_col in value_columns:
                            units_value = self.extract_number_from_cell(
                                sheet, row, val_col
                            )
                            if units_value > 0:
                                units = int(units_value)
                                break
                        if units is not None:
                            break
            if units is not None:
                break

        # 2. Dynamic search for Total square footage of all project structures
        # Search for label in multiple columns, extract value from AG or L
        for row in range(1, sheet.max_row + 1):
            for col in label_columns:
                cell_value = sheet.cell(row=row, column=col).value
                if (
                    cell_value
                    and "total square footage of all project structures"
                    in str(cell_value).lower()
                ):
                    # Try AG first, then L
                    for val_col in value_columns:
                        sf_value = self.extract_number_from_cell(sheet, row, val_col)
                        if sf_value > 0:
                            sf = sf_value
                            break
                    if sf is not None:
                        break
            if sf is not None:
                break

        return units, sf

    def parse(self) -> ApplicationData:
        """Parse the entire application and return ApplicationData"""
        self.load_workbook()

        app_data = ApplicationData()
        app_data.file_path = self.file_path
        app_data.application_name = Path(self.file_path).stem

        # Extract financing data
        app_data.construction_interest_fees = self.extract_construction_interest_fees()
        app_data.permanent_financing = self.extract_permanent_financing()

        # Extract additional data
        app_data.new_construction_total = self.extract_new_construction_total()
        app_data.total_units, app_data.total_square_feet = self.extract_units_and_sf()

        # Validate
        app_data.validate()

        return app_data
